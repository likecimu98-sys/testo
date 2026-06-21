from __future__ import annotations

import csv
import json
import re
import shutil
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from pypdf import PdfReader


OUT_IMAGE_DIR = Path("assets/visual/painting")
OUT_DATA = Path("visualPaintingData.generated.js")
OUT_MARKDOWN = Path("painting_characteristics.md")
OUT_CSV = Path("painting_characteristics.csv")
OUT_XLSX = Path("painting_characteristics.xlsx")
OUT_HTML = Path("painting_characteristics.html")
SOURCE_NAME = "папа.pdf"


ENTRIES = [
    {
        "title": "Спас Вседержитель",
        "creator": "Феофан Грек",
        "dateText": "1378 г.",
        "years": [1378],
        "century": "XIV век",
        "halfCentury": "вторая половина XIV века",
        "style": "древнерусская монументальная живопись; византийская традиция",
        "kind": "фреска",
        "description": "Роспись купола церкви Спаса Преображения на Ильине улице в Новгороде.",
        "imageRef": [2, 1],
        "sourcePages": [2],
    },
    {
        "title": "Богоматерь Донская",
        "creator": "Феофан Грек",
        "dateText": "1392 г.",
        "years": [1392],
        "century": "XIV век",
        "halfCentury": "вторая половина XIV века",
        "style": "древнерусская иконопись; византийская традиция",
        "kind": "икона",
        "description": "Современное местонахождение: Государственная Третьяковская галерея, Москва.",
        "imageRef": [3, 1],
        "sourcePages": [3, 4],
    },
    {
        "title": "Троица Ветхозаветная",
        "creator": "Андрей Рублев",
        "dateText": "1425-1427 гг.",
        "years": [1425, 1427],
        "century": "XV век",
        "halfCentury": "первая половина XV века",
        "style": "древнерусская иконопись; московская школа",
        "kind": "икона",
        "description": "Храмовая икона иконостаса Троице-Сергиева монастыря; до 2023 г. находилась в Государственной Третьяковской галерее.",
        "imageRef": [5, 1],
        "sourcePages": [5, 6],
    },
    {
        "title": "Христос Вседержитель",
        "creator": "Дионисий",
        "dateText": "1502-1503 гг.",
        "years": [1502, 1503],
        "century": "XVI век",
        "halfCentury": "первая половина XVI века",
        "style": "древнерусская монументальная живопись; московская школа",
        "kind": "фреска",
        "description": "Роспись подкупольного пространства собора Рождества Богородицы Ферапонтова монастыря.",
        "imageRef": [6, 1],
        "sourcePages": [6],
    },
    {
        "title": "Спас Нерукотворный",
        "creator": "Симон Ушаков",
        "dateText": "1658 г.",
        "years": [1658],
        "century": "XVII век",
        "halfCentury": "вторая половина XVII века",
        "style": "иконопись Оружейной палаты",
        "kind": "икона",
        "description": "Образец перехода от средневековой иконописи к более объемной живописной манере XVII века.",
        "imageRef": [7, 1],
        "sourcePages": [7],
    },
    {
        "title": "Похвала иконе «Богоматерь Владимирская» (Древо государства Российского)",
        "creator": "Симон Ушаков",
        "dateText": "1663 г.",
        "years": [1663],
        "century": "XVII век",
        "halfCentury": "вторая половина XVII века",
        "style": "иконопись Оружейной палаты; аллегорическая композиция",
        "kind": "икона",
        "description": "Аллегорическое прославление Владимирской иконы Богоматери и российской государственности.",
        "imageRef": [8, 1],
        "sourcePages": [8, 9],
    },
    {
        "title": "Царь Алексей Михайлович в зрелые годы",
        "creator": "неизвестный художник",
        "dateText": "середина XVII в.",
        "years": [],
        "century": "XVII век",
        "halfCentury": "середина XVII века",
        "style": "парсуна",
        "kind": "парсуна",
        "description": "Ранний светский портрет, сохраняющий черты иконописной традиции.",
        "imageRef": [10, 1],
        "sourcePages": [10, 11],
    },
    {
        "title": "Коронационный портрет Екатерины II",
        "creator": "Федор Рокотов",
        "dateText": "1763 г.",
        "years": [1763],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "парадный портрет; рококо и классицизм",
        "kind": "портрет",
        "description": "Парадный императорский портрет Екатерины II.",
        "imageRef": [12, 1],
        "sourcePages": [12],
    },
    {
        "title": "Владимир и Рогнеда",
        "creator": "Антон Лосенко",
        "dateText": "1770 г.",
        "years": [1770],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "классицизм; историческая живопись",
        "kind": "историческая картина",
        "description": "Картина на сюжет древнерусской истории.",
        "imageRef": [13, 1],
        "sourcePages": [13],
    },
    {
        "title": "Прощание Гектора с Андромахой",
        "creator": "Антон Лосенко",
        "dateText": "1773 г.",
        "years": [1773],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "классицизм; историко-мифологическая живопись",
        "kind": "историческая картина",
        "description": "Картина на античный сюжет из истории Троянской войны.",
        "imageRef": [14, 1],
        "sourcePages": [14],
    },
    {
        "title": "Портрет Прокофия Демидова",
        "creator": "Дмитрий Левицкий",
        "dateText": "1773 г.",
        "years": [1773],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "парадный портрет; классицизм",
        "kind": "портрет",
        "description": "Парадный портрет промышленника и мецената Прокофия Демидова.",
        "imageRef": [15, 1],
        "sourcePages": [15, 16],
    },
    {
        "title": "Портрет неизвестной крестьянки в русском костюме",
        "creator": "Иван Аргунов",
        "dateText": "1784 г.",
        "years": [1784],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "портретная живопись XVIII века; реалистическая традиция",
        "kind": "портрет",
        "description": "Известный образ крестьянки в русском национальном костюме.",
        "imageRef": [17, 1],
        "sourcePages": [17, 18],
    },
    {
        "title": "Портрет Марии Лопухиной",
        "creator": "Владимир Боровиковский",
        "dateText": "1797 г.",
        "years": [1797],
        "century": "XVIII век",
        "halfCentury": "вторая половина XVIII века",
        "style": "сентиментализм; камерный портрет",
        "kind": "портрет",
        "description": "Камерный женский портрет конца XVIII века.",
        "imageRef": [19, 1],
        "sourcePages": [19],
    },
    {
        "title": "Портрет Александра Пушкина",
        "creator": "Орест Кипренский",
        "dateText": "1827 г.",
        "years": [1827],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века",
        "style": "романтизм; портрет",
        "kind": "портрет",
        "description": "Один из самых известных портретов А. С. Пушкина.",
        "imageRef": [21, 1],
        "sourcePages": [21],
    },
    {
        "title": "Портрет Александра Пушкина",
        "creator": "Василий Тропинин",
        "dateText": "1827 г.",
        "years": [1827],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века",
        "style": "реалистический портрет; романтизм",
        "kind": "портрет",
        "description": "Портрет Пушкина в домашнем образе.",
        "imageRef": [22, 1],
        "sourcePages": [22],
    },
    {
        "title": "Последний день Помпеи",
        "creator": "Карл Брюллов",
        "dateText": "1833 г.",
        "years": [1833],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века",
        "style": "романтизм; академизм; историческая живопись",
        "kind": "историческая картина",
        "description": "Крупная историческая композиция на античный сюжет.",
        "imageRef": [23, 1],
        "sourcePages": [23],
    },
    {
        "title": "Явление Христа народу",
        "creator": "Александр Иванов",
        "dateText": "1837-1857 гг.",
        "years": [1837, 1857],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века; завершена во второй половине",
        "style": "академизм; историко-религиозная живопись",
        "kind": "историческая картина",
        "description": "Монументальное полотно на евангельский сюжет, создавалось около двадцати лет.",
        "imageRef": [24, 1],
        "sourcePages": [24],
    },
    {
        "title": "Свежий кавалер (Утро чиновника, получившего первый крестик)",
        "creator": "Павел Федотов",
        "dateText": "1846 г.",
        "years": [1846],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века",
        "style": "критический реализм; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Сатирическая бытовая сцена о чиновничьей среде.",
        "imageRef": [25, 1],
        "sourcePages": [25],
    },
    {
        "title": "Сватовство майора",
        "creator": "Павел Федотов",
        "dateText": "1848 г.",
        "years": [1848],
        "century": "XIX век",
        "halfCentury": "первая половина XIX века",
        "style": "критический реализм; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Сатирическая сцена из купеческо-дворянского быта.",
        "imageRef": [26, 1],
        "sourcePages": [26],
    },
    {
        "title": "Сельский крестный ход на Пасхе",
        "creator": "Василий Перов",
        "dateText": "1861 г.",
        "years": [1861],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "критический реализм; бытовой жанр; передвижники",
        "kind": "жанровая картина",
        "description": "Критическое изображение сельского религиозного быта.",
        "imageRef": [27, 1],
        "sourcePages": [27],
    },
    {
        "title": "Тройка. Ученики-мастеровые везут воду",
        "creator": "Василий Перов",
        "dateText": "1866 г.",
        "years": [1866],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "критический реализм; бытовой жанр; передвижники",
        "kind": "жанровая картина",
        "description": "Социально-критическая картина о детском труде.",
        "imageRef": [28, 1],
        "sourcePages": [28],
    },
    {
        "title": "Петр I допрашивает царевича Алексея Петровича в Петергофе",
        "creator": "Николай Ге",
        "dateText": "1871 г.",
        "years": [1871],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реализм; историческая живопись",
        "kind": "историческая картина",
        "description": "Драматическая сцена конфликта Петра I и царевича Алексея.",
        "imageRef": [29, 1],
        "sourcePages": [29],
    },
    {
        "title": "Грачи прилетели",
        "creator": "Алексей Саврасов",
        "dateText": "1871 г.",
        "years": [1871],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "лирический пейзаж; передвижники",
        "kind": "пейзаж",
        "description": "Классический русский лирический пейзаж.",
        "imageRef": [30, 1],
        "sourcePages": [30, 31],
    },
    {
        "title": "Земство обедает",
        "creator": "Григорий Мясоедов",
        "dateText": "1872 г.",
        "years": [1872],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "критический реализм; бытовой жанр; передвижники",
        "kind": "жанровая картина",
        "description": "Социально-критическое изображение земской действительности.",
        "imageRef": [31, 1],
        "sourcePages": [31],
    },
    {
        "title": "Апофеоз войны",
        "creator": "Василий Верещагин",
        "dateText": "1871 г.",
        "years": [1871],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "батальный жанр; реализм; антивоенная живопись",
        "kind": "батальная картина",
        "description": "Антивоенное полотно с символическим образом последствий войны.",
        "imageRef": [32, 1],
        "sourcePages": [32],
    },
    {
        "title": "Двери Тимура (Тамерлана)",
        "creator": "Василий Верещагин",
        "dateText": "1872 г.",
        "years": [1872],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "батальный и исторический жанр; реализм",
        "kind": "историческая картина",
        "description": "Картина восточной серии Верещагина.",
        "imageRef": [33, 1],
        "sourcePages": [33],
    },
    {
        "title": "Бурлаки на Волге",
        "creator": "Илья Репин",
        "dateText": "1872-1873 гг.",
        "years": [1872, 1873],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реализм; передвижники; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Социально-критическая картина о тяжелом труде бурлаков.",
        "imageRef": [34, 1],
        "sourcePages": [34],
    },
    {
        "title": "Крестный ход в Курской губернии",
        "creator": "Илья Репин",
        "dateText": "1880-1883 гг.",
        "years": [1880, 1883],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реализм; передвижники; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Многофигурная сцена русской общественной жизни.",
        "imageRef": [35, 1],
        "sourcePages": [35],
    },
    {
        "title": "Запорожцы",
        "creator": "Илья Репин",
        "dateText": "1880-1891 гг.",
        "years": [1880, 1891],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм",
        "kind": "историческая картина",
        "description": "Историческое полотно о запорожских казаках.",
        "imageRef": [36, 1],
        "sourcePages": [36],
    },
    {
        "title": "Иван Грозный и сын его Иван 16 ноября 1581 года",
        "creator": "Илья Репин",
        "dateText": "1883 г.",
        "years": [1883],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм; психологизм",
        "kind": "историческая картина",
        "description": "Психологически напряженная историческая сцена.",
        "imageRef": [37, 1],
        "sourcePages": [37],
    },
    {
        "title": "Курсистка",
        "creator": "Николай Ярошенко",
        "dateText": "1883 г.",
        "years": [1883],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реализм; портретный жанр; передвижники",
        "kind": "портрет",
        "description": "Образ женщины-студентки как символа новых общественных явлений.",
        "imageRef": [38, 1],
        "sourcePages": [38, 39],
    },
    {
        "title": "Арест пропагандиста",
        "creator": "Илья Репин",
        "dateText": "1880-1889 гг.",
        "years": [1880, 1889],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реализм; революционная тема",
        "kind": "жанровая картина",
        "description": "Сцена ареста революционного пропагандиста.",
        "imageRef": [39, 2],
        "sourcePages": [39],
    },
    {
        "title": "Утро стрелецкой казни",
        "creator": "Василий Суриков",
        "dateText": "1881 г.",
        "years": [1881],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм; передвижники",
        "kind": "историческая картина",
        "description": "Историческое полотно о стрелецком бунте и эпохе Петра I.",
        "imageRef": [40, 1],
        "sourcePages": [40],
    },
    {
        "title": "Меншиков в Берёзове",
        "creator": "Василий Суриков",
        "dateText": "1883 г.",
        "years": [1883],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм; передвижники",
        "kind": "историческая картина",
        "description": "Картина о ссылке А. Д. Меншикова.",
        "imageRef": [41, 1],
        "sourcePages": [41],
    },
    {
        "title": "Боярыня Морозова",
        "creator": "Василий Суриков",
        "dateText": "1887 г.",
        "years": [1887],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм; передвижники",
        "kind": "историческая картина",
        "description": "Историческое полотно о церковном расколе XVII века.",
        "imageRef": [42, 1],
        "sourcePages": [42],
    },
    {
        "title": "Покорение Ермаком Сибири",
        "creator": "Василий Суриков",
        "dateText": "1895 г.",
        "years": [1895],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм",
        "kind": "историческая картина",
        "description": "Историческое полотно о походе Ермака.",
        "imageRef": [43, 1],
        "sourcePages": [43],
    },
    {
        "title": "Переход Суворова через Альпы",
        "creator": "Василий Суриков",
        "dateText": "1899 г.",
        "years": [1899],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "историческая живопись; реализм",
        "kind": "историческая картина",
        "description": "Картина о Швейцарском походе А. В. Суворова.",
        "imageRef": [44, 1],
        "sourcePages": [44, 45],
    },
    {
        "title": "Портрет Льва Николаевича Толстого",
        "creator": "Иван Крамской",
        "dateText": "1873 г.",
        "years": [1873],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реалистический портрет; передвижники",
        "kind": "портрет",
        "description": "Портрет Л. Н. Толстого.",
        "imageRef": [46, 1],
        "sourcePages": [46],
    },
    {
        "title": "Портрет Николая Некрасова",
        "creator": "Иван Крамской",
        "dateText": "1877 г.",
        "years": [1877],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реалистический портрет; передвижники",
        "kind": "портрет",
        "description": "Портрет Н. А. Некрасова.",
        "imageRef": [47, 1],
        "sourcePages": [47, 48],
    },
    {
        "title": "Портрет Михаила Салтыкова-Щедрина",
        "creator": "Иван Крамской",
        "dateText": "1879 г.",
        "years": [1879],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "реалистический портрет; передвижники",
        "kind": "портрет",
        "description": "Портрет М. Е. Салтыкова-Щедрина.",
        "imageRef": [49, 1],
        "sourcePages": [49, 50],
    },
    {
        "title": "Аленушка",
        "creator": "Виктор Васнецов",
        "dateText": "1881 г.",
        "years": [1881],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "сказочно-былинный жанр; неорусский стиль",
        "kind": "сказочная картина",
        "description": "Картина по мотивам русской сказочной традиции.",
        "imageRef": [51, 1],
        "sourcePages": [51, 52],
    },
    {
        "title": "Богатыри",
        "creator": "Виктор Васнецов",
        "dateText": "1881-1898 гг.",
        "years": [1881, 1898],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "сказочно-былинный жанр; неорусский стиль",
        "kind": "былинная картина",
        "description": "Образ трех богатырей русской былинной традиции.",
        "imageRef": [52, 2],
        "sourcePages": [52],
    },
    {
        "title": "Девочка с персиками",
        "creator": "Валентин Серов",
        "dateText": "1887 г.",
        "years": [1887],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "русский импрессионизм; портрет",
        "kind": "портрет",
        "description": "Портрет Веры Мамонтовой, один из символов русского импрессионизма.",
        "imageRef": [53, 1],
        "sourcePages": [53],
    },
    {
        "title": "Демон сидящий",
        "creator": "Михаил Врубель",
        "dateText": "1890 г.",
        "years": [1890],
        "century": "XIX век",
        "halfCentury": "вторая половина XIX века",
        "style": "символизм; модерн",
        "kind": "символистская картина",
        "description": "Один из центральных образов символистской живописи Врубеля.",
        "imageRef": [54, 1],
        "sourcePages": [54],
    },
    {
        "title": "Царевна-Лебедь",
        "creator": "Михаил Врубель",
        "dateText": "1900 г.",
        "years": [1900],
        "century": "рубеж XIX-XX вв.",
        "halfCentury": "рубеж XIX-XX веков",
        "style": "символизм; модерн",
        "kind": "сказочная картина",
        "description": "Сказочно-символистский образ на рубеже XIX-XX веков.",
        "imageRef": [55, 1],
        "sourcePages": [55, 56],
    },
    {
        "title": "Солдатушки, бравы ребятушки, где же ваша слава?",
        "creator": "Валентин Серов",
        "dateText": "1905 г.",
        "years": [1905],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "политическая сатира; модернистская графика",
        "kind": "графика",
        "description": "Антивоенная и антисамодержавная графическая работа периода революции 1905 года.",
        "imageRef": [56, 2],
        "sourcePages": [56],
    },
    {
        "title": "Похищение Европы",
        "creator": "Валентин Серов",
        "dateText": "1910 г.",
        "years": [1910],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "модерн; античная тема",
        "kind": "мифологическая картина",
        "description": "Модернистская трактовка античного мифа.",
        "imageRef": [57, 1],
        "sourcePages": [57],
    },
    {
        "title": "Купание красного коня",
        "creator": "Кузьма Петров-Водкин",
        "dateText": "1912 г.",
        "years": [1912],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "символизм; модерн",
        "kind": "символистская картина",
        "description": "Символический образ предреволюционной эпохи.",
        "imageRef": [58, 1],
        "sourcePages": [58],
    },
    {
        "title": "Композиция VII",
        "creator": "Василий Кандинский",
        "dateText": "1913 г.",
        "years": [1913],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "абстракционизм; авангард",
        "kind": "абстрактная картина",
        "description": "Один из ключевых памятников русского авангарда.",
        "imageRef": [59, 1],
        "sourcePages": [59],
    },
    {
        "title": "Черный супрематический квадрат",
        "creator": "Казимир Малевич",
        "dateText": "1915 г.",
        "years": [1915],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "супрематизм; авангард",
        "kind": "авангардная картина",
        "description": "Ключевой символ супрематизма и русского авангарда.",
        "imageRef": [60, 1],
        "sourcePages": [60],
    },
    {
        "title": "Большевик",
        "creator": "Борис Кустодиев",
        "dateText": "1920 г.",
        "years": [1920],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "раннесоветская живопись; революционный романтизм",
        "kind": "историко-революционная картина",
        "description": "Монументальный образ революционной массы и большевика.",
        "imageRef": [61, 1],
        "sourcePages": [61],
    },
    {
        "title": "Тачанка",
        "creator": "Митрофан Греков",
        "dateText": "1925 г.",
        "years": [1925],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "советский батальный жанр",
        "kind": "батальная картина",
        "description": "Образ Гражданской войны в советской батальной живописи.",
        "imageRef": [62, 1],
        "sourcePages": [62],
    },
    {
        "title": "Новая Москва",
        "creator": "Юрий Пименов",
        "dateText": "1937 г.",
        "years": [1937],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "социалистический реализм; городской пейзаж",
        "kind": "жанровая картина",
        "description": "Оптимистический образ советской столицы 1930-х годов.",
        "imageRef": [63, 1],
        "sourcePages": [63],
    },
    {
        "title": "Родина-мать зовет",
        "creator": "Ираклий Тоидзе",
        "dateText": "1941 г.",
        "years": [1941],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "агитационный плакат; социалистический реализм",
        "kind": "плакат",
        "description": "Один из главных советских плакатов начала Великой Отечественной войны.",
        "imageRef": [64, 1],
        "sourcePages": [64, 65],
    },
    {
        "title": "Беспощадно разгромим и уничтожим врага",
        "creator": "Кукрыниксы",
        "dateText": "1941 г.",
        "years": [1941],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "агитационный плакат; политическая сатира",
        "kind": "плакат",
        "description": "Военный плакат первых месяцев Великой Отечественной войны.",
        "imageRef": [66, 1],
        "sourcePages": [66, 67],
    },
    {
        "title": "Пьём воду родного Днепра",
        "creator": "Виктор Иванов",
        "dateText": "1943 г.",
        "years": [1943],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "военный агитационный плакат",
        "kind": "плакат",
        "description": "Плакат, связанный с освобождением территории СССР в годы Великой Отечественной войны.",
        "imageRef": [68, 1],
        "sourcePages": [68, 69],
    },
    {
        "title": "Оборона Севастополя",
        "creator": "Александр Дейнека",
        "dateText": "1942 г.",
        "years": [1942],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "батальная живопись; социалистический реализм",
        "kind": "батальная картина",
        "description": "Героико-батальная картина периода Великой Отечественной войны.",
        "imageRef": [69, 2],
        "sourcePages": [69],
    },
    {
        "title": "Утро нашей Родины",
        "creator": "Федор Шурпин",
        "dateText": "1948 г.",
        "years": [1948],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "социалистический реализм; культ личности",
        "kind": "портрет-картина",
        "description": "Полотно периода позднего сталинизма.",
        "imageRef": [70, 1],
        "sourcePages": [70],
    },
    {
        "title": "Прибыл на каникулы",
        "creator": "Федор Решетников",
        "dateText": "1948 г.",
        "years": [1948],
        "century": "XX век",
        "halfCentury": "первая половина XX века",
        "style": "социалистический реализм; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Бытовая сцена из советской школьной жизни.",
        "imageRef": [71, 1],
        "sourcePages": [71, 72],
    },
    {
        "title": "Опять двойка",
        "creator": "Федор Решетников",
        "dateText": "1952 г.",
        "years": [1952],
        "century": "XX век",
        "halfCentury": "вторая половина XX века",
        "style": "социалистический реализм; бытовой жанр",
        "kind": "жанровая картина",
        "description": "Одна из самых узнаваемых советских бытовых картин.",
        "imageRef": [72, 2],
        "sourcePages": [72],
    },
    {
        "title": "Строители Братской ГЭС",
        "creator": "Виктор Попков",
        "dateText": "1960-1961 гг.",
        "years": [1960, 1961],
        "century": "XX век",
        "halfCentury": "вторая половина XX века",
        "style": "суровый стиль; позднесоветская живопись",
        "kind": "жанровая картина",
        "description": "Памятник позднесоветской живописи эпохи оттепели.",
        "imageRef": [73, 1],
        "sourcePages": [73],
    },
]


TRANSLIT = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "j",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "c",
    "ч": "ch",
    "ш": "sh",
    "щ": "shch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


def slugify(text: str) -> str:
    out = []
    for ch in text.lower():
        out.append(TRANSLIT.get(ch, ch if ch.isascii() else "-"))
    slug = re.sub(r"[^a-z0-9]+", "-", "".join(out)).strip("-")
    return slug or "painting"


def find_source_pdf() -> Path:
    desktop = Path(r"C:\Users\01\Desktop")
    candidates = sorted(desktop.glob("*.pdf"), key=lambda p: p.stat().st_size, reverse=True)
    for candidate in candidates:
        try:
            reader = PdfReader(str(candidate))
            text = reader.pages[0].extract_text() or ""
        except Exception:
            continue
        if "живописи" in text and "ЕГЭ" in text:
            return candidate
    raise FileNotFoundError("Не найден PDF со списком живописи на рабочем столе.")


def period_for(entry: dict) -> str:
    century = entry["century"]
    if century in {"XIV век", "XV век", "XVI век", "XVII век"}:
        return "early"
    if century == "XVIII век":
        return "18th"
    if century == "XIX век" or century == "рубеж XIX-XX вв.":
        return "19th"
    return "20th"


def row_characteristic(entry: dict) -> str:
    parts = [
        f"Название: {entry['title']}",
        f"Автор: {entry['creator']}",
        f"Дата: {entry['dateText']}",
        f"Половина века: {entry['halfCentury']}",
        f"Стиль/направление: {entry['style']}",
        f"Тип: {entry['kind']}",
    ]
    if entry.get("description"):
        parts.append(entry["description"])
    return "; ".join(parts)


def drill_facts(entry: dict) -> list[dict]:
    return [
        {
            "type": "title",
            "label": "название",
            "question": "Как называется этот памятник живописи?",
            "answer": entry["title"],
        },
        {
            "type": "creator",
            "label": "автор",
            "question": "Кто автор этого произведения?",
            "answer": entry["creator"],
        },
        {
            "type": "halfCentury",
            "label": "половина века",
            "question": "К какой половине века относится произведение?",
            "answer": entry["halfCentury"],
        },
        {
            "type": "style",
            "label": "стиль",
            "question": "Какой стиль или направление лучше всего подходит?",
            "answer": entry["style"],
        },
    ]


def build_data_entry(entry: dict, image_path: str, slug: str) -> dict:
    traits = [
        f"Автор: {entry['creator']}.",
        f"Дата: {entry['dateText']}; {entry['halfCentury']}.",
        f"Стиль/направление: {entry['style']}.",
        f"Тип: {entry['kind']}.",
    ]
    if entry.get("description"):
        traits.append(entry["description"])
    return {
        "id": slug,
        "title": entry["title"],
        "type": "painting",
        "period": period_for(entry),
        "century": entry["century"],
        "halfCentury": entry["halfCentury"],
        "years": entry["years"],
        "dateText": entry["dateText"],
        "locations": [],
        "styles": [entry["style"]],
        "creators": [entry["creator"]],
        "rulers": [],
        "events": [],
        "description": entry["description"],
        "traits": traits,
        "fullCharacteristic": row_characteristic(entry),
        "importantFacts": traits,
        "drillFacts": drill_facts(entry),
        "images": [image_path],
        "mainImage": image_path,
        "source": {
            "file": SOURCE_NAME,
            "pages": entry["sourcePages"],
            "imageRef": entry["imageRef"],
        },
        "reviewed": False,
    }


def extract_images(reader: PdfReader) -> dict[tuple[int, int], bytes]:
    page_images: dict[tuple[int, int], bytes] = {}
    for page_no, page in enumerate(reader.pages, start=1):
        for img_no, image in enumerate(page.images, start=1):
            page_images[(page_no, img_no)] = image.data
    return page_images


def write_image(data: bytes, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    img = Image.open(BytesIO(data)).convert("RGB")
    img.save(destination, "JPEG", quality=95, optimize=True)


def write_markdown(rows: list[dict]) -> None:
    lines = [
        "# Живопись для визуального тренажера ЕГЭ",
        "",
        "| № | Изображение | Название | Автор | Дата | Половина века | Стиль/направление | Тип |",
        "|---:|---|---|---|---|---|---|---|",
    ]
    for idx, row in enumerate(rows, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(idx),
                    f"![{row['title']}]({row['imagePath']})",
                    row["title"].replace("|", "\\|"),
                    row["creator"].replace("|", "\\|"),
                    row["dateText"].replace("|", "\\|"),
                    row["halfCentury"].replace("|", "\\|"),
                    row["style"].replace("|", "\\|"),
                    row["kind"].replace("|", "\\|"),
                ]
            )
            + " |"
        )
    OUT_MARKDOWN.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_csv(rows: list[dict]) -> None:
    fields = [
        "№",
        "Название",
        "Автор",
        "Дата",
        "Век",
        "Половина века",
        "Стиль/направление",
        "Тип",
        "Ключевые характеристики",
        "Страницы PDF",
        "Файл изображения",
    ]
    with OUT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for idx, row in enumerate(rows, start=1):
            writer.writerow(
                {
                    "№": idx,
                    "Название": row["title"],
                    "Автор": row["creator"],
                    "Дата": row["dateText"],
                    "Век": row["century"],
                    "Половина века": row["halfCentury"],
                    "Стиль/направление": row["style"],
                    "Тип": row["kind"],
                    "Ключевые характеристики": row_characteristic(row),
                    "Страницы PDF": ", ".join(map(str, row["sourcePages"])),
                    "Файл изображения": row["imagePath"],
                }
            )


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Живопись"
    headers = [
        "№",
        "Изображение",
        "Название",
        "Автор",
        "Дата",
        "Век",
        "Половина века",
        "Стиль/направление",
        "Тип",
        "Ключевые характеристики",
        "Страницы PDF",
        "Файл изображения",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [6, 20, 38, 24, 16, 16, 30, 34, 20, 70, 14, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for idx, row in enumerate(rows, start=2):
        ws.row_dimensions[idx].height = 92
        values = [
            idx - 1,
            "",
            row["title"],
            row["creator"],
            row["dateText"],
            row["century"],
            row["halfCentury"],
            row["style"],
            row["kind"],
            row_characteristic(row),
            ", ".join(map(str, row["sourcePages"])),
            row["imagePath"],
        ]
        ws.append(values)
        for col in range(1, len(headers) + 1):
            ws.cell(idx, col).alignment = Alignment(vertical="top", wrap_text=True)
        img = XlsxImage(row["imagePath"])
        img.width = 110
        img.height = 82
        ws.add_image(img, f"B{idx}")
    wb.save(OUT_XLSX)


def write_html(rows: list[dict]) -> None:
    cards = []
    for idx, row in enumerate(rows, start=1):
        cards.append(
            f"""
        <article class="card">
          <img src="{row['imagePath']}" alt="{row['title']}">
          <div class="body">
            <div class="num">{idx:02d}</div>
            <h2>{row['title']}</h2>
            <dl>
              <dt>Автор</dt><dd>{row['creator']}</dd>
              <dt>Дата</dt><dd>{row['dateText']}</dd>
              <dt>Половина века</dt><dd>{row['halfCentury']}</dd>
              <dt>Стиль</dt><dd>{row['style']}</dd>
              <dt>Тип</dt><dd>{row['kind']}</dd>
            </dl>
            <p>{row['description']}</p>
          </div>
        </article>"""
        )
    html = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Живопись ЕГЭ: визуальный каталог</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #172033;
      --muted: #5e6878;
      --line: #d8dee8;
      --paper: #f7f8fb;
      --card: #ffffff;
      --accent: #315b8a;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, sans-serif;
      background: var(--paper);
      color: var(--ink);
    }}
    header {{
      padding: 28px 24px 18px;
      border-bottom: 1px solid var(--line);
      background: #fff;
    }}
    h1 {{
      margin: 0;
      font-size: 28px;
      line-height: 1.15;
    }}
    .meta {{
      margin-top: 8px;
      color: var(--muted);
      font-size: 14px;
    }}
    main {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
      gap: 14px;
      padding: 18px;
    }}
    .card {{
      display: grid;
      grid-template-rows: 220px auto;
      min-width: 0;
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
    }}
    .card img {{
      width: 100%;
      height: 100%;
      object-fit: contain;
      background: #eef1f5;
      padding: 8px;
    }}
    .body {{
      padding: 12px;
      position: relative;
    }}
    .num {{
      color: var(--accent);
      font-weight: 700;
      font-size: 12px;
      letter-spacing: .08em;
    }}
    h2 {{
      margin: 4px 0 10px;
      font-size: 17px;
      line-height: 1.2;
    }}
    dl {{
      display: grid;
      grid-template-columns: 94px 1fr;
      gap: 6px 8px;
      margin: 0;
      font-size: 13px;
    }}
    dt {{
      color: var(--muted);
      font-weight: 700;
    }}
    dd {{
      margin: 0;
    }}
    p {{
      margin: 10px 0 0;
      font-size: 13px;
      color: var(--muted);
      line-height: 1.4;
    }}
  </style>
</head>
<body>
  <header>
    <h1>Живопись ЕГЭ: визуальный каталог</h1>
    <div class="meta">61 объект из файла {SOURCE_NAME}. Поля: название, автор, дата, половина века, стиль и тип.</div>
  </header>
  <main>
    {''.join(cards)}
  </main>
</body>
</html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")


def main() -> None:
    pdf = find_source_pdf()
    reader = PdfReader(str(pdf))
    page_images = extract_images(reader)
    if OUT_IMAGE_DIR.exists():
        shutil.rmtree(OUT_IMAGE_DIR)
    OUT_IMAGE_DIR.mkdir(parents=True, exist_ok=True)

    seen: dict[str, int] = {}
    rows = []
    data_entries = []
    for entry in ENTRIES:
        base = slugify(f"{entry['title']} {entry['creator']}")
        seen[base] = seen.get(base, 0) + 1
        slug = base if seen[base] == 1 else f"{base}-{seen[base]}"
        image_key = tuple(entry["imageRef"])
        if image_key not in page_images:
            raise KeyError(f"Нет изображения {image_key} для {entry['title']}")
        image_path = OUT_IMAGE_DIR / f"{slug}.jpg"
        write_image(page_images[image_key], image_path)
        image_rel = image_path.as_posix()
        row = {**entry, "id": slug, "imagePath": image_rel}
        rows.append(row)
        data_entries.append(build_data_entry(entry, image_rel, slug))

    OUT_DATA.write_text(
        "// Generated by scripts/extract_visual_painting_pdf.py\n"
        "// Source PDF is user-provided; style fields are normalized for EGE visual training.\n"
        "window.visualPaintingData = "
        + json.dumps(data_entries, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    write_markdown(rows)
    write_csv(rows)
    write_xlsx(rows)
    write_html(rows)
    print(f"source: {pdf}")
    print(f"entries: {len(rows)}")
    print(f"images: {OUT_IMAGE_DIR}")
    print(f"data: {OUT_DATA}")
    print(f"markdown: {OUT_MARKDOWN}")
    print(f"csv: {OUT_CSV}")
    print(f"xlsx: {OUT_XLSX}")
    print(f"html: {OUT_HTML}")


if __name__ == "__main__":
    main()
